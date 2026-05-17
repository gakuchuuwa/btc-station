"""
把 VectorBT dynamic_runner 的回测结果导出为 xlsx。
5个Sheet对标参考文件格式：表现 / 交易分析 / 风险调整后的表现 / 交易清单 / 属性
无装饰颜色，简洁黑白格式。
"""
import io
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BOLD = Font(bold=True, name="Calibri", size=10)
_NORM = Font(name="Calibri", size=10)
_HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
_THIN_SIDE = Side(style="thin", color="BFBFBF")
_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)


def _v(metrics: dict, key: str, default=None):
    val = metrics.get(key, default)
    if val is None:
        return None
    try:
        f = float(val)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return val


def _set_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _hdr(ws, row: int, values: list):
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = _BOLD
        c.fill = _HDR_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24


def _row(ws, row: int, label: str, values: list):
    c0 = ws.cell(row=row, column=1, value=label)
    c0.font = _NORM
    c0.border = _BORDER
    for ci, v in enumerate(values, 2):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = _NORM
        c.border = _BORDER
        c.alignment = Alignment(horizontal="right")


# ── Sheet 1：表现 ─────────────────────────────────────────────────────────────

def _sheet_performance(wb: Workbook, m: dict):
    """对标 TV「表现」33 行"""
    ws = wb.create_sheet("表现")
    _set_width(ws, 1, 32)
    for ci in range(2, 8):
        _set_width(ws, ci, 14)

    _hdr(ws, 1, ["", "全部 USDT", "全部 %", "多头 USDT", "多头 %", "空头 USDT", "空头 %"])

    init = _v(m, "initial_capital", 10000) or 10000
    np_l = _v(m, "net_profit_long",  0) or 0
    np_s = _v(m, "net_profit_short", 0) or 0
    rp_l = _v(m, "return_pct_long")
    rp_s = _v(m, "return_pct_short")
    gpl  = _v(m, "gross_profit_long",  0) or 0
    gll  = _v(m, "gross_loss_long",    0) or 0
    gps  = _v(m, "gross_profit_short", 0) or 0
    gls  = _v(m, "gross_loss_short",   0) or 0
    gpl_pct = round(gpl / init * 100, 4) if init else None
    gll_pct = round(gll / init * 100, 4) if init else None
    gps_pct = round(gps / init * 100, 4) if init else None
    gls_pct = round(gls / init * 100, 4) if init else None
    gpa_pct = round((_v(m,"gross_profit_abs") or 0) / init * 100, 4) if init else None
    gla_pct = round((_v(m,"gross_loss_abs")   or 0) / init * 100, 4) if init else None
    np_pct  = _v(m,"total_return_pct")
    bm_abs  = _v(m,"benchmark_return_abs")
    bm_pct  = _v(m,"benchmark_return_pct")

    def neg(v):
        return round(-abs(v), 2) if v is not None else None

    rows = [
        ("初始资本",         [init, None, None, None, None, None]),
        ("未实现盈亏",       [_v(m,"open_trade_pnl") or 0, 0, None, None, None, None]),
        ("净利润",           [_v(m,"net_profit_abs"), np_pct, np_l, rp_l, np_s, rp_s]),
        ("毛利润",           [_v(m,"gross_profit_abs"), gpa_pct, gpl, gpl_pct, gps, gps_pct]),
        ("毛亏损",           [neg(_v(m,"gross_loss_abs")), neg(gla_pct), neg(gll), neg(gll_pct), neg(gls), neg(gls_pct)]),
        ("预期收益",         [_v(m,"expectancy_abs"), None, _v(m,"expectancy_long"), None, _v(m,"expectancy_short"), None]),
        ("已支付佣金",       [_v(m,"commission_paid"), None, None, None, None, None]),
        ("买入和持有回报",   [bm_abs, bm_pct, None, None, None, None]),
        ("买入并持有收益率", [None, bm_pct, None, None, None, None]),
        ("策略表现优异",     [_v(m,"strategy_outperformance"), None, None, None, None, None]),
        ("最大合同持有量",   [_v(m,"max_position_size"), None, None, None, None, None]),
        ("年化收益率(CAGR)", [None, _v(m,"cagr_pct"), None, _v(m,"cagr_pct_long"), None, _v(m,"cagr_pct_short")]),
        ("初始资本回报率",   [None, np_pct, None, rp_l, None, rp_s]),

        ("平均净值涨幅",     [_v(m,"avg_runup_abs"), _v(m,"avg_runup_pct"), None, None, None, None]),
        ("平均涨幅持续时间", [f"{_v(m,'avg_runup_duration_days') or 0}天" if _v(m,"avg_runup_duration_days") else None,
                              None, None, None, None, None]),
        ("最大净值涨幅",     [_v(m,"max_runup_abs"), _v(m,"max_runup_pct"), None, None, None, None]),
        ("最大涨幅(intrabar)", [_v(m,"max_runup_intrabar_abs"), _v(m,"max_runup_intrabar_pct"), None, None, None, None]),

        ("平均净值回撤",     [None, _v(m,"avg_drawdown_pct"), None, None, None, None]),
        ("平均回撤持续",     [f"{_v(m,'avg_drawdown_duration_days') or 0}天" if _v(m,"avg_drawdown_duration_days") else None,
                              None, None, None, None, None]),
        ("最大净值回撤",     [None, neg(_v(m,"max_drawdown_pct")), None, None, None, None]),
        ("最大回撤持续",     [f"{_v(m,'max_drawdown_duration_days') or 0}天" if _v(m,"max_drawdown_duration_days") else None,
                              None, None, None, None, None]),
        ("最大回撤(intrabar)", [None, neg(_v(m,"max_drawdown_pct")), None, None, None, None]),
        ("最大回撤占初始资本%(intrabar)", [None, neg(_v(m,"ftmo_drawdown_pct")), None, None, None, None]),
        ("最大回撤波谷剩余利润", [_v(m,"max_dd_profit_at_trough"), None, None, None, None, None]),

        ("净利润占最大亏损%", [None, _v(m,"net_profit_over_max_loss_pct"), None, None, None, None]),
        ("最大盈利占总盈利%", [None, _v(m,"max_win_over_gross_profit_pct"), None, None, None, None]),
        ("最大亏损占总亏损%", [None, _v(m,"max_loss_over_gross_loss_pct"), None, None, None, None]),
    ]
    for ri, (label, vals) in enumerate(rows, 2):
        _row(ws, ri, label, vals)


# ── Sheet 2：交易分析 ─────────────────────────────────────────────────────────

def _sheet_trade_analysis(wb: Workbook, m: dict):
    """对标 TV「交易分析」18 行,7 列(多/空分组)"""
    ws = wb.create_sheet("交易分析")
    _set_width(ws, 1, 22)
    for ci in range(2, 8):
        _set_width(ws, ci, 14)

    _hdr(ws, 1, ["", "全部 USDT", "全部 %", "多头 USDT", "多头 %", "空头 USDT", "空头 %"])

    tl  = _v(m, "total_trades_long",  0) or 0
    ts  = _v(m, "total_trades_short", 0) or 0
    wl  = _v(m, "win_trades_long",    0) or 0
    ws_ = _v(m, "win_trades_short",   0) or 0
    ll_ = _v(m, "loss_trades_long",   0) or 0
    ls_ = _v(m, "loss_trades_short",  0) or 0
    long_wr  = round(wl  / tl  * 100, 2) if tl  else None
    short_wr = round(ws_ / ts * 100, 2) if ts  else None

    def neg(v):
        return round(-abs(v), 2) if v is not None else None

    rows = [
        ("总未平仓交易",      [0, None, 0, None, 0, None]),
        ("总交易",            [_v(m,"total_trades"), None, tl, None, ts, None]),
        ("盈利交易",          [_v(m,"win_trades"),   None, wl, None, ws_, None]),
        ("亏损交易",          [_v(m,"loss_trades"),  None, ll_, None, ls_, None]),
        ("均等交易",          [0, None, 0, None, 0, None]),
        ("获利百分比",        [None, _v(m,"win_rate_pct"), None, long_wr, None, short_wr]),
        ("平均盈亏",          [_v(m,"expectancy_abs"), None, _v(m,"expectancy_long"), None, _v(m,"expectancy_short"), None]),
        ("平均盈利交易",      [_v(m,"avg_win_abs"), _v(m,"avg_win_pct"), _v(m,"avg_win_long_abs"), None, _v(m,"avg_win_short_abs"), None]),
        ("平均亏损交易",      [_v(m,"avg_loss_abs"), _v(m,"avg_loss_pct"), _v(m,"avg_loss_long_abs"), None, _v(m,"avg_loss_short_abs"), None]),
        ("平均胜率/平均负率", [_v(m,"payoff_ratio"), None, _v(m,"payoff_long"), None, _v(m,"payoff_short"), None]),
        ("最大盈利交易",      [_v(m,"max_win_abs"), None, _v(m,"max_win_long_abs"), None, _v(m,"max_win_short_abs"), None]),
        ("最大盈利交易百分比", [None, _v(m,"max_win_pct"), None, None, None, None]),
        ("最大亏损交易",      [_v(m,"max_loss_abs"), None, _v(m,"max_loss_long_abs"), None, _v(m,"max_loss_short_abs"), None]),
        ("最大亏损交易百分比", [None, _v(m,"max_loss_pct"), None, None, None, None]),
        ("交易的平均#K线数",  [_v(m,"avg_bars_all"), None, _v(m,"avg_bars_long"), None, _v(m,"avg_bars_short"), None]),
        ("盈利交易的平均#K线数", [_v(m,"avg_bars_win"), None, None, None, None, None]),
        ("亏损交易的平均#K线数", [_v(m,"avg_bars_loss"), None, None, None, None, None]),
    ]
    for ri, (label, vals) in enumerate(rows, 2):
        _row(ws, ri, label, vals)


# ── Sheet 3：风险调整后的表现 ─────────────────────────────────────────────────

def _sheet_risk(wb: Workbook, m: dict):
    """对标 TV「风险调整后的表现」5 行,7 列(盈利因子多空分组)"""
    ws = wb.create_sheet("风险调整后的表现")
    _set_width(ws, 1, 22)
    for ci in range(2, 8):
        _set_width(ws, ci, 14)

    _hdr(ws, 1, ["", "全部 USDT", "全部 %", "多头 USDT", "多头 %", "空头 USDT", "空头 %"])

    # 多/空盈利因子 = gross_profit / gross_loss
    pf_l = pf_s = None
    try:
        gpl = _v(m, "gross_profit_long",  0) or 0
        gll = _v(m, "gross_loss_long",    0) or 0
        gps = _v(m, "gross_profit_short", 0) or 0
        gls = _v(m, "gross_loss_short",   0) or 0
        if gll: pf_l = round(gpl / gll, 3)
        if gls: pf_s = round(gps / gls, 3)
    except Exception:
        pass

    rows = [
        ("夏普比率",   [_v(m,"sharpe"),  None, None, None, None, None]),
        ("Sortino比率",[_v(m,"sortino"), None, None, None, None, None]),
        ("盈利因子",   [_v(m,"profit_factor"), None, pf_l, None, pf_s, None]),
        ("追加保证金", [0, None, 0, None, 0, None]),
    ]
    for ri, (label, vals) in enumerate(rows, 2):
        _row(ws, ri, label, vals)


# ── Sheet 4：交易清单 ─────────────────────────────────────────────────────────

def _sheet_trades(wb: Workbook, trades: list, initial_capital: float = 10000.0):
    ws = wb.create_sheet("交易清单")

    headers = [
        "交易 #", "类型", "日期和时间", "信号",
        "价格 USDT", "大小（数量）", "大小（价值）",
        "净损益 USDT", "净损益 %",
        "有利波动 USDT", "有利波动 %",
        "不利波动 USDT", "不利波动 %",
        "累计P&L USDT", "累计P&L %",
    ]
    widths = [8, 12, 20, 28, 14, 13, 14, 13, 10, 13, 10, 13, 10, 14, 10]
    for ci, w in enumerate(widths, 1):
        _set_width(ws, ci, w)

    _hdr(ws, 1, headers)
    ws.freeze_panes = "A2"

    def _parse_ts(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v
        s = str(v)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s[:19], fmt)
            except ValueError:
                pass
        return None

    cum_pnl = 0.0
    for i, t in enumerate(trades, 1):
        pnl_abs = round(float(t.get("PnL", 0)), 2)
        pnl_pct = round(float(t.get("Return", 0)) * 100, 2)
        cum_pnl += pnl_abs
        cum_pct  = round(cum_pnl / initial_capital * 100, 2)
        direction = str(t.get("Direction", "Long")).replace("Direction.", "")
        dir_label = "多头" if direction.lower() == "long" else "空头"
        entry_dt = _parse_ts(t.get("Entry Timestamp"))
        exit_dt  = _parse_ts(t.get("Exit Timestamp"))
        entry_px = round(float(t.get("Avg Entry Price", t.get("Entry Price", 0))), 2)
        exit_px  = round(float(t.get("Avg Exit Price",  t.get("Exit Price",  0))), 2)
        size_qty = round(float(t.get("Size", 0)), 6)
        # 大小（价值）= 数量 × 入场价（持仓名义价值，与 TV 一致）
        size_val = round(size_qty * entry_px, 4)
        # 入场信号（首入="S2 Long-P1"、加仓="Scale-in Long 1"）+ 出场信号（"Trailing SL" 等）
        entry_sig = str(t.get("Signal", ""))
        exit_sig  = str(t.get("ExitSignal", ""))

        base = 2 + (i - 1) * 2

        def write_row(row_num, type_label, dt, px, signal_text):
            vals = [i, type_label, dt, signal_text, px, size_qty, size_val,
                    pnl_abs, pnl_pct, None, None, None, None,
                    round(cum_pnl, 2), cum_pct]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=ci, value=v)
                c.font = _NORM
                c.border = _BORDER
                c.alignment = Alignment(horizontal="right" if ci > 2 else "left")
                if ci == 3 and isinstance(v, datetime):
                    c.number_format = "YYYY/MM/DD HH:MM"

        write_row(base,     f"{dir_label}出场", exit_dt,  exit_px,  exit_sig)
        write_row(base + 1, f"{dir_label}进场", entry_dt, entry_px, entry_sig)


# ── Sheet 5：属性 ─────────────────────────────────────────────────────────────

def _sheet_attrs(wb: Workbook, m: dict, strategy_name: str, timeframe: str):
    """对标 TV「属性」71 行 — 完整暴露回测元信息 + 所有策略参数"""
    ws = wb.create_sheet("属性")
    _set_width(ws, 1, 30)
    _set_width(ws, 2, 36)

    _hdr(ws, 1, ["name", "value"])

    tf_map = {
        "1m":"1分钟","5m":"5分钟","15m":"15分钟","30m":"30分钟",
        "1h":"1小时","2h":"2小时","4h":"4小时","6h":"6小时","12h":"12小时",
        "1d":"1天","1w":"1周",
    }
    start = str(m.get("backtest_start", ""))
    end   = str(m.get("backtest_end",   ""))
    init  = _v(m, "initial_capital", 10000)

    # 1. 系统属性(类似 TV 的"商品代码/时间周期/Tick"等)
    attrs = [
        ("交易范围",   f"{start} — {end}"),
        ("回测范围",   f"{start} — {end}"),
        ("商品代码",   "OKX:BTCUSDT.P"),
        ("时间周期",   tf_map.get(timeframe, timeframe)),
        ("货币",       "USDT"),
        ("Tick大小",   0.1),
        ("初始资金",   init),
        ("数据来源",   "OKX 公共 API"),
        ("策略名称",   strategy_name),
        ("导出时间",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    # 2. 所有策略参数(从 raw_parameters 完整复制,把内部参数过滤掉)
    raw_params = m.get("raw_parameters") or {}
    skip_prefix = ('_', )  # 跳过 _strategy_name / _timeframe 等内部字段
    skip_keys   = {"initial_capital", "fixed_qty"}  # 已经在系统属性显示
    if isinstance(raw_params, dict):
        for k, v in raw_params.items():
            if k in skip_keys or k.startswith(skip_prefix): continue
            # bool 转中文
            if isinstance(v, bool):
                v = "打开" if v else "关闭"
            attrs.append((str(k), v))

    for ri, (k, v) in enumerate(attrs, 2):
        ws.cell(row=ri, column=1, value=k).font = _NORM
        ws.cell(row=ri, column=2, value=v).font  = _NORM
        ws.cell(row=ri, column=1).border = _BORDER
        ws.cell(row=ri, column=2).border = _BORDER


# ── 主入口 ────────────────────────────────────────────────────────────────────

def vectorbt_to_xlsx(
    raw: dict,
    strategy_name: str = "自定义策略",
    timeframe: str = "4h",
) -> bytes:
    m      = raw.get("metrics", {})
    trades = raw.get("trades", [])
    init   = float(m.get("initial_capital", 10000) or 10000)

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_performance(wb, m)
    _sheet_trade_analysis(wb, m)
    _sheet_risk(wb, m)
    _sheet_trades(wb, trades, init)
    _sheet_attrs(wb, m, strategy_name, timeframe)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
